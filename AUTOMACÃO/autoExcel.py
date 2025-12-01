from openpyxl import load_workbook

arquivo = load_workbook("AUTOMACÃO/Alunos.xlsx")

### Ver as abas
print(arquivo.sheetnames)

### Pegar a aba ativa
aba_atual = arquivo.active
print(aba_atual)

### Selecionar uma aba especifica
aba_alunos = arquivo["Planilha1"] ### passa o neme dela no colchetes ([])
print(aba_alunos)

### Selecionar células
valor_a1 = aba_alunos["A1"].value ### se precisar confere com : print(aba_alunos["A1"].value)
valor_b1 = aba_alunos.cell(row=1, column=2).value ### a a mesma coisa de cima (linha 17)
print(valor_b1)

aba_alunos.cell(row=1, column=2).value = "Prova 1"

arquivo.save("Alunos2.xlsx") ### salva como um novo arquivo. Se quiser pode : ("Aluno.xlsx") que substitui o antigo

### Ultima linha
print(aba_alunos.max_row)
print(len(aba_alunos["A"])) ### Outro jeito de fazer
